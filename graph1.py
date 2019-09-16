import os
import statistics
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import (
    RadarChart,
    ScatterChart,
    Reference,
    Series,
)


class GraphIt:

    def __init__(self):
        self.header_line = 0
        self.last_line = 0
        self.rot_degrees_col = 2
        self.atten_col = 3
        self.rssi_col = 16
        self.rx_col = 17
        self.thru_col = 6 
        self.status_col = 12


    def delete_file(self, fil):
        if os.path.exists(fil):
            os.remove(fil)


    def read_all_lines(self, file_name):
        # develop a list of all lines in the .csv file
        content = []
        with open(file_name) as f:
            content = f.readlines()
        return content


    def parse_header(self, lines):
        # determine column number of each applicable header
        # determine header_line and last_line
        entries = []
        headers = {}
        line_no = 0
        for line in lines:
            if self.header_line > 0 and line.strip() == '':
                self.last_line = line_no
                break
            entries = line.split(',')
            for entry in entries:
                entry = entry.strip()
                if 'Test Run' in entry:
                    self.header_line = line_no
                    for intN in range(len(entries)):
                        headers.update({entries[intN]: intN})
            line_no += 1
        # determine applicable header column numbers
        for key, value in headers.items():
            if 'Position' in key:
                self.rot_degrees_col = value
            elif 'Attenuation' in key:
                self.atten_col = value
            elif 'Data Rssi' in key:
                self.rssi_col = value
            elif 'Rx Rate' in key:
                self.rx_col = value
            elif ' Traffic Pair 1 Throughput [Mbps]' in key:
                self.thru_col = value
            elif 'Test Status' in key:
                self.status_col = value

    def parse_rotational_angles(self, lines):
        # determine applicable degrees of rotation
        positions = []
        for intN in range(self.header_line + 1, self.last_line):
            line = lines[intN]
            entries = line.split(',')
            if len(entries) < self.rot_degrees_col:
                continue
            entry = entries[self.rot_degrees_col]
            entry = entry.strip()
            if entry.isdigit():
                positions.append(entry)
        positions = list(dict.fromkeys(positions))
        return positions


    def parse_attenuations(self, lines):
        # determine applicable degrees of rotation
        attenuations = []
        for intN in range(self.header_line + 1, self.last_line):
            line = lines[intN]
            entries = line.split(',')
            if len(entries) < self.atten_col:
                continue
            entry = entries[self.atten_col]
            entry = entry.strip()
            if entry.isdigit():
                attenuations.append(entry)
        attenuations = list(dict.fromkeys(attenuations))
        return attenuations


    def is_float(self, num):
        try:
            numb = float(num)
            return True
        except:
            return False
        
    
    def build_the_damned_thing(self, lines, angles, attens):
        graph_list = []
        for angle in angles:
            for atten in attens:
                rssi_tot = 0
                rssi_ct = 0
                thru_tot = 0
                thru_ct = 0
                for intN in range(self.header_line + 1, self.last_line):
                    line = lines[intN]
                    entries = line.split(',') 
                    if '(avg)' in (entries)[self.status_col]:
                        if (entries[self.rot_degrees_col]).strip() == angle and (entries[self.atten_col]).strip() == atten:
                            rssi_entry = entries[self.rssi_col]
                            rssi_entry = rssi_entry.strip()
                            if self.is_float(rssi_entry):
                                rssi_tot += float(rssi_entry)
                                rssi_ct +=1
                            thru_entry = entries[self.thru_col]
                            thru_entry = thru_entry.strip()
                            if self.is_float(thru_entry):
                                thru_tot += float(thru_entry)
                                thru_ct +=1
                rssi_avgsour = 0
                if rssi_ct > 0:
                    rssi_avg = round(float(rssi_tot/rssi_ct), 1)
                thru_avg = 0
                if thru_ct > 0:
                    thru_avg = round(float(thru_tot/thru_ct), 1)
                line_dict = {"rot_angle": angle, "atten": atten, "rssi_avg": rssi_avg, "thru_avg": thru_avg}
                graph_list.append(line_dict)
        return graph_list


    def rssi_values(self, lines):
        rssi_values = []
        for intN in range(self.header_line + 1, self.last_line):
            line = lines[intN]
            entries = line.split(',')
            rssi_entry = entries[self.rssi_col]
            rssi_entry = rssi_entry.strip()
            if self.is_float(rssi_entry):
                rssi = float(rssi_entry)
                rssi_values.append(rssi)
        rssi_values = list(dict.fromkeys(rssi_values))
        return rssi_values


    def throughputs_by_rssi(self, lines, rssi):
        throughput_values = []
        for intN in range(self.header_line + 1, self.last_line):
            line = lines[intN]
            entries = line.split(',')
            rssi_entry = entries[self.rssi_col]
            rssi_entry = rssi_entry.strip()
            if self.is_float(rssi_entry):
                this_rssi = float(rssi_entry)
                if this_rssi == rssi:
                    throughput_entry = entries[self.thru_col]
                    througput_entry = throughput_entry.strip()
                    if self.is_float(throughput_entry):
                        throughput = float(throughput_entry)
                        throughput_values.append(throughput)
        throughput_values = list(dict.fromkeys(throughput_values))
        return throughput_values


    def standard_deviations(self, rssi, throughs):
        if len(throughs) > 1:
            mean_thru = statistics.mean(throughs)
            mean_thru = round(mean_thru, 2)
            std_dev = statistics.stdev(throughs)
            std_dev = round(std_dev, 2)
        elif len(throughs) == 1:
            mean_thru = throughs[0]
            std_dev = 0
        else:
            return None
        throughput_rssi_spread = {'rssi': rssi, 'count': len(throughs), 'mean_thru': mean_thru, 'std_dev': std_dev}
        return throughput_rssi_spread


    def testing123(self, lines):
        for intN in range (15):
            line = lines[intN]
            entries = line.split(',')
            if 'Test Run' in entries[0]:
                hello = entries[self.thru_col]
                goodbye = entries[self.rssi_col]
                parking_space = 1


    
if __name__ == "__main__":
    Grph = GraphIt()
    fil = r'C:\VSCode\Graphs\randeep_onebox\RvRvO 2G US.csv'

    Grph.delete_file('example.xlsx')
    all_lines = Grph.read_all_lines(fil)
    Grph.parse_header(all_lines)
    angles = Grph.parse_rotational_angles(all_lines)
    attens = Grph.parse_attenuations(all_lines)                
    data_set = Grph.build_the_damned_thing(all_lines, angles, attens)

    Grph.testing123(all_lines)    

    rssi_list = Grph.rssi_values(all_lines)
    rssi_list.sort()
    list_of_stats = []
    for rssi_entry in rssi_list:
        list_of_thrus = Grph.throughputs_by_rssi(all_lines, rssi_entry)
        stats =  Grph.standard_deviations(rssi_entry, list_of_thrus)
        if stats != None:
            list_of_stats.append(stats)
            parking_space = 2
    parking_space = 3

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Radar Graph"
    # Row & Column Headers
    # Rows must be rotational angle
    # Data Sets 
    intN = 0
    for r in range(2, 2 + len(angles)):
        ws.cell(row = r, column = 1, value = str(angles[intN]) + ' deg')
        intN +=1
    intN = 0
    for c in range(2, 2 + len(attens)):
        ws.cell(row = 1, column = c, value = str(attens[intN]) + 'dB' )
        intN +=1

    '''line_dict = {"rot_angle": angle, "atten": atten, "rssi_avg": rssi_avg, "thru_avg": thru_avg}
    graph_list.append(line_dict)'''        
    # Build Rate vs Atten vs Orientation
    for intColumn in range(len(attens)):
        c = 2 + intColumn
        this_atten = attens[intColumn]
        for intRow in range(len(angles)):
            r = 2 + intRow
            this_angle = angles[intRow]
            cell_value = 0
            for datum in data_set:
                if datum['rot_angle'] == this_angle and datum['atten'] == this_atten:
                    cell_value = datum['thru_avg']
                    break
            ws.cell(row = r, column = c, value = cell_value)

    chart = RadarChart()
    #chart.type = "filled"
    labels = Reference(ws, min_col=1, min_row=2, max_row=len(angles) + 1)
    data = Reference(ws, min_col=2, max_col=len(attens) + 1, min_row= 1, max_row=len(angles) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    for intN in range(len(attens)):
        chart.series[intN].graphicalProperties.line.width = 1    
    #chart.style = 26
    chart.title = "Rate vs Attenuation vs Orientation"
    chart.width = chart.width * 1.5
    chart.height = chart.height * 2
    chart.y_axis.majorUnit = 5
    #chart.y_axis.delete = True

    ws.add_chart(chart, "A17")


    # Data Sets 
    intN = 0
    for r in range(2, 2 + len(angles)):
        ws.cell(row = r, column = len(attens) + 3, value = str(angles[intN]) + ' deg')
        intN +=1
    intN = 0
    for c in range(2, 2 + len(attens)):
        ws.cell(row = 1, column = c + len(attens) + 2, value = str(attens[intN]) + 'dB' )
        intN +=1
    # Build Rate vs RSSI vs Orientation
    for intColumn in range(len(attens)):
        c = len(attens) + 4 + intColumn
        this_atten = attens[intColumn]
        for intRow in range(len(angles)):
            r = 2 + intRow
            this_angle = angles[intRow]
            cell_value = 0
            for datum in data_set:
                if datum['rot_angle'] == this_angle and datum['atten'] == this_atten:
                    cell_value = datum['rssi_avg']
                    break
            ws.cell(row = r, column = c, value = cell_value)

    # Get min/max rssi values
    max_rssi = -1000
    min_rssi = 1000
    for datum in data_set:
        this_rssi = float(datum['rssi_avg'])
        if this_rssi > max_rssi:
            max_rssi = this_rssi
        if this_rssi < min_rssi:
            min_rssi = this_rssi

    chart2 = RadarChart()
    data_start = len(attens) + 4
    labels = Reference(ws, min_col=1, min_row=2, max_row=len(angles) + 1)
    data = Reference(ws, min_col=data_start, max_col=data_start + len(attens) - 1, min_row= 1, max_row=len(angles) + 1)
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(labels)
    for intN in range(len(attens)):
        chart2.series[intN].graphicalProperties.line.width = 1    
    chart2.title = "RSSI vs Attenuation vs Orientation"
    chart2.width = chart.width 
    chart2.height = chart.height
    chart2.y_axis.majorUnit = 5
    chart2.y_axis.scaling.max  = max_rssi + 5
    start_ascii = 65 + len(attens) + 2
    start_column = chr(start_ascii) 
    start_cell = start_column + "17"
    ws.add_chart(chart2, start_cell)


    # Build RSSI vs THRU chart
    '''line_dict = {"rot_angle": angle, "atten": atten, "rssi_avg": rssi_avg, "thru_avg": thru_avg}
    graph_list.append(line_dict)'''
    # Sort the list
    # sorted(lis, key = lambda i: i['age'])
    hello = sorted(data_set, key = lambda i: i['rssi_avg'])
    # develop a sorted list of all unique rssis
    all_rssis = []
    for datum in hello:
        all_rssis.append(datum['rssi_avg'])
    all_rssis = list(dict.fromkeys(all_rssis))


    start_row = len(angles) + 10
    ws.cell(row=start_row - 1, column=1, value="RSSI")
    fin_col = 0
    for intRSSI in range(len(all_rssis)):
        this_rssi = all_rssis[intRSSI]
        r = start_row + intRSSI
        c = 2
        for datum in hello:
            rssi_val = datum['rssi_avg']
            if rssi_val == this_rssi:
                ws.cell(row=r, column=1, value = rssi_val)
                thru_val = datum['thru_avg']
                ws.cell(row=r, column=c, value=thru_val)
                c += 1
                if c > fin_col:
                    fin_col = c
                            
    chart3 = ScatterChart()
    chart3.title = "Throughput vs RSSI"
    #chart3.style = 13
    chart3.x_axis.title = 'RSSI'
    chart3.y_axis.title = 'THROUGHPUT'
    xvalues = Reference(ws, min_col=1, min_row=start_row, max_row=start_row + len(all_rssis))

    for i in range(2, fin_col):
        values = Reference(ws, min_col=i, min_row=start_row, max_row=start_row + len(all_rssis))
        series = Series(values, xvalues, title_from_data=True)
        series.marker = openpyxl.chart.marker.Marker('circle')
        series.graphicalProperties.line.noFill=True
        chart3.series.append(series)
    chart3.x_axis.majorUnit = 1
    chart3.x_axis.scaling.max  = int(max_rssi) + 1
    chart3.x_axis.scaling.min = int(min_rssi) - 1
    chart3.width = chart3.width * 3



    '''for intN in range(len(chart3.series)):
        chart3.series[intN].graphicalProperties.line.width = 1'''
    ws.add_chart(chart3, "A50")

    wb.save('example.xlsx')

    os.startfile('example.xlsx')




    parking = 1


    
